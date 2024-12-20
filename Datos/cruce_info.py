import win32com.client as win32
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook

from Datos.pasos import Pasos


class GuardarDatos(Pasos):

    def __init__(self):
        #Inicializa una instancia de Excel para trabajar con archivos de Excel.
        
        #self.excel = win32.gencache.EnsureDispatch('Excel.Application')
        #self.excel.Visible = False  # Cambiar a True para ver el proceso en Excel.
        self.path_ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.path_Reportes= os.path.join(self.path_,'Reportes')
        self.path_ResultadoArchivo= os.path.join(self.path_,'Resultado', 'Adulto_mayor.xlsx')

    """ def cargar_archivos(self):
        #Carga los archivos necesarios para el procesamiento.
        columnas_Pagos=['Fecha','Ciudad','Zona','Cédula Cajero/Vendedor','Cajero/Vendedor',
            'Canal', 'Reclama titular?', 'Empresa', 'Producto','Cod. Oficina', 'Oficina',
            'Cod. Sitio ','Sitio de venta','Tipo Doc. Titular', 'Identificación Titular',  'Titular',
            'Tipo Doc. Autorizado', 'Identificación Autorizado', 'Autorizado','Periodo', 'Valor Reportado', 
            'Valor Pagado', 'Valor Redondeo', 'Grupo Pago','Tipo Pago','Tipo Subsidio', 'Titular Comfama', 
            'Titular Comfama']
        self.reporte_pagos = pd.read_excel('reportes/reportePagos.xlsx',usecols=columnas_Pagos)
        print(self.reporte_pagos)

        columnas_Usuarios=['Identificacion','Cargo','Estado']
        self.reporte_usuarios = pd.read_excel('reportes/reporteUsuarios.xlsx',usecols=columnas_Usuarios)
        print(self.reporte_usuarios)

         #Procesa los datos cargados, separa fecha y hora, y selecciona las columnas de interés.
        # Separar la columna FechaHora en Fecha y Hora

        self.reporte_pagos.columns = self.reporte_pagos.columns.str.strip()
        self.reporte_pagos[['Fecha', 'Hora']] = self.reporte_pagos['Fecha'].str.split(' ', expand=True)
        print("Reporte Pagos después de separar Fecha y Hora:",self.reporte_pagos)
        print(self.reporte_pagos[['Fecha', 'Hora']].head()) """
         
    """ def procesar_datos(self):
        #Procesa los datos cargados, separa fecha y hora, y selecciona las columnas de interés.
        # Separar la columna FechaHora en Fecha y Hora

        self.reporte_pagos.columns = self.reporte_pagos.columns.str.strip()
        self.reporte_pagos[['Fecha', 'Hora']] = self.reporte_pagos['Fecha'].str.split(' ', expand=True)
        # Mapeo de columnas según tu especificación

        self.mapeo_columnas = {
            'A': 'Fecha',  'B': 'Hora',  'C': 'Ciudad',  'D': 'Zona','E': 'Cédula Cajero/Vendedor',  'F': 'Cajero/Vendedor',
            'H': 'Canal','I': 'Reclama titular?','J': 'Empresa',  'K': 'Producto','L': 'Cod. Oficina',  'M': 'Oficina',
            'N': 'Cod. Sitio','O': 'Sitio de venta','P': 'Tipo Doc. Titular','Q': 'Identificación Titular',  'R': 'Titular',
            'S': 'Tipo Doc. Autorizado','T': 'Identificación Autorizado','U': 'Autorizado','X': 'Periodo','Y': 'Valor Reportado', 
            'Z': 'Valor Pagado','AA': 'Valor Redondeo','AB': 'Grupo Pago','AC': 'Tipo Pago','AD': 'Tipo Subsidio','AE': 'Titular Comfama', 
            'AF': 'Titular Comfama'}
        # Mapeo de columnas de reporte_usuarios
        self.mapeo_columnas_usuarios = {
            'G':'Cargo','AG':'Estado',} """

    """ def convertir_columna_a_indice(self, columna):
        #Convierte una columna en formato letra (A, B, AA) a un índice numérico.
        resultado = 0
        for char in columna:
            resultado = resultado * 26 + (ord(char.upper()) - ord('A') + 1)
        return resultado """
    
    def __lista_archivo_final(self):
        print(self.path_ResultadoArchivo)
        df = pd.read_excel(self.path_ResultadoArchivo)
        
         # Crear una lista con los nombres de las columnas
        titulos_lista = list(df.columns)
        print("Lista de títulos:", titulos_lista)
        return titulos_lista

    def __cargar_archivos_usuarios1(self):
        path_reporte_usuarios = os.path.join(self.path_Reportes, 'reporteUsuarios.xlsx' )
        path_reporte_usuarios_m = os.path.join(self.path_Reportes, 'reporteUsuarios_modificado.xlsx' )
        #Se carga el archivo de usuarios
        wb_reporte_usuarios = load_workbook(path_reporte_usuarios)
        hoja = wb_reporte_usuarios.active
        # Lista de columnas a eliminar
        columnas_a_eliminar = [2, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28]
        # Eliminar las columnas desde el final hacia el principio para evitar problemas con el cambio de índice
        for col in sorted(columnas_a_eliminar, reverse=True):
            print(f"Eliminando columna usuarios: {col}")
            hoja.delete_cols(col)
        #wb_reporte_usuarios.save(path_reporte_usuarios_m)
        return hoja
    
    def __cargar_archivos_usuarios(self):
        columnas_Usuarios=['Identificacion','Cargo','Estado','Oficina']
        reporte_usuarios = pd.read_excel('reportes/reporteUsuarios.xlsx',usecols=columnas_Usuarios)
        print(reporte_usuarios)
        return reporte_usuarios

    def __cargar_archivos_pagos(self):
        path_reporte_pagos = os.path.join(self.path_Reportes,'reportePagos.xlsx' )
        path_reporte_pagos_m = os.path.join(self.path_Reportes,'reportePagos_modificado.xlsx' )
        #Se carga el archivo de pagos
        wb_reporte_pagos = load_workbook(path_reporte_pagos)
        # Selecciona la hoja activa
        hoja = wb_reporte_pagos.active
        #Se crea lista para eliminar columnas 
        columnas_a_eliminar =[2,4,8]
        # Eliminar las columnas desde el final hacia el principio para evitar problemas con el cambio de índice
        for col in sorted(columnas_a_eliminar, reverse=True):
            print(f"Eliminando columna pagos: {col}")
            hoja.delete_cols(col)
        # Insertar una nueva columna para la hora
        # Inserta una nueva columna en la posición 2 (después de la columna A)
        hoja.insert_cols(2)  
        hoja.cell(row=1, column=2, value="Hora")
        # Procesar la columna A (fecha y hora) y dividir los datos
        # Ajusta las filas según tu archivo
        for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row, min_col=1, max_col=1):
            # Celda de la columna A (suponiendo que ahí están las fechas y horas) 
            celda = fila[0] 
            # Verificar que la celda no esté vacía
            if celda.value:  
                try:
                    # Convertir el valor a datetime
                    fecha_hora = datetime.strptime(celda.value, '%d/%m/%Y %H:%M:%S')
                    fecha = fecha_hora.date()
                    hora = fecha_hora.time()
                    # Sobreescribir la columna A con solo la fecha
                     # Columna A para la fecha
                    hoja.cell(row=celda.row, column=1, value=str(fecha)) 
                    # Escribir la hora en la nueva columna (columna B)
                    hoja.cell(row=celda.row, column=2, value=str(hora))  # Columna B para la hora
                except Exception as e:
                    print(f"Error al convertir la celda en la fila {celda.row}: {celda.value} {e}")

        # Guardar los cambios en un nuevo archivo
        #wb_reporte_pagos.save(path_reporte_pagos_m)
        return hoja

    def cruzar_archivos(self):
        hora_inicio = datetime.now()
        print(hora_inicio)
        hoja_reporte_pagos = self.__cargar_archivos_pagos()
        reporte_usuarios = self.__cargar_archivos_usuarios()
        titulos_lista = self.__lista_archivo_final()

        # Crear un nuevo Workbook para el resultado
        wb_resultado = load_workbook(self.path_ResultadoArchivo)
        hoja_resultado = wb_resultado['DATA']

        #hoja_pagos = wb_reporte_pagos.active

        # Transferir los datos de wb_reporte_pagos al nuevo workbook
        fila = 2
        for row_num, fila in enumerate(hoja_reporte_pagos.iter_rows(min_row=2, values_only=True), 2):
            # Llenar las columnas según los títulos de titulos_lista
            identificacion = fila[19]
            print(f"identificacion a llenar: {identificacion}")
            
            cargo_autorizado = self.__buscar_usuarios_autorizado(reporte_usuarios, identificacion)
            i = 0

            col_  = 0
            #for cell in hoja_reporte_pagos[1]:

            for col_num, titulo in enumerate(titulos_lista):
        
            # Suponiendo que los títulos están en la fila 1
            #for cell in hoja_reporte_pagos[1]:
                """ if titulo == cell.value or titulo == "Hora":
                    print(f"error: Titulo: {titulo} cell: {cell}")
                    hoja_resultado.cell(row=row_num, column=col_num+1, value=fila[col_num])
                    print(f"Celda a llenar: {titulo} valor: {fila[col_num]}")
                    col_ +=1
                    break """
            
            #if titulo == hoja_reporte_pagos.columns[titulo]  # Asegurarse de que el título esté presente en wb_reporte_pagos
                
                if titulo == "Cargo":
                    identificacion = fila[4]
                    cargo = self.__buscar_usuarios_cajero(reporte_usuarios, identificacion)
                    if cargo:
                        print(f"Celda a llenar: {titulo} valor: '{cargo}'")
                        hoja_resultado.cell(row=row_num, column=col_num + 1, value=cargo)
                    """ else:
                        hoja_resultado.cell(row=row_num, column=col_, value='') """

                    

                elif titulo == "Cargo autorizado" or titulo == "Oficina1" or titulo == "Estado":
                    
                    if cargo_autorizado is not None:
                        hoja_resultado.cell(row=row_num, column=col_num + 1, value=cargo_autorizado[i])
                        print(f"Celda a llenar: {titulo} valor: '{cargo_autorizado[i]}'")
                    """ else:
                        hoja_resultado.cell(row=row_num, column=col_, value='')
                        print(f"Celda a llenar: {titulo} valor: 'vacio'") """
                    i += 1

                    
                else:
                    columna_encontrada = None
                    for col in hoja_reporte_pagos.iter_cols(min_row=1, max_row=1):  # Iterar solo sobre la primera fila
                        if col[0].value == titulo:
                            columna_encontrada = col[0].column  # Obtener el índice de la columna
                            
                            hoja_resultado.cell(row=row_num, column=col_num+1, value=fila[col_])
                            print(f"Celda a llenar: {titulo} valor: {fila[col_]}")
                            col_ += 1
                            break
                    
                #col_index = hoja_reporte_pagos.columns.index(col_num) + 1
                # Copiar los datos de la columna correspondiente de wb_reporte_pagos
                """ if col_ == 29:
                    # Guardar el nuevo workbook con los datos combinados
                    wb_resultado.save(self.path_ResultadoArchivo)
                    hora_final = datetime.now()
                    print(f"hora inicio {hora_inicio}. hora final {hora_final}")
                    break
            if col_ == 29:
                break """
                """ hoja_resultado.cell(row=row_num, column=col_, value=fila[col_ - 1])
                print(f"Celda a llenar: {titulo} valor: {fila[col_num]}") """

        # Guardar el nuevo workbook con los datos combinados
        wb_resultado.save(self.path_ResultadoArchivo)
        hora_final = datetime.now()
        print(f"hora inicio {hora_inicio}. hora final {hora_final}")

    def __buscar_usuarios_cajero(self, reporte_usuarios, num_id):
        print(type(reporte_usuarios))
        print(type(reporte_usuarios['Identificacion']))
        print(reporte_usuarios.columns)
        cruce_cajero = reporte_usuarios[reporte_usuarios['Identificacion'].astype(str).str.strip() == str(num_id).strip()]
        #cruce_cajero = reporte_usuarios[reporte_usuarios['Identificacion'] == num_id]
        print(cruce_cajero)
        if not cruce_cajero.empty:
            cargo = cruce_cajero['Cargo'].iloc[0].strip()
            print(cargo)
        else:
            cargo = None
        return cargo
        

    def __buscar_usuarios_autorizado(self, reporte_usuarios, num_id):
        usuario_lista = []
        cruce_usuario =  reporte_usuarios[reporte_usuarios['Identificacion'].astype(str).str.strip() == str(num_id).strip()]

        if not cruce_usuario.empty:
            cargo = cruce_usuario['Cargo'].iloc[0].strip()
            oficina = cruce_usuario['Oficina'].iloc[0].strip()
            estado = cruce_usuario['Estado'].iloc[0].strip()
            usuario_lista = [cargo, oficina, estado]
        else:
            usuario_lista = None
        return usuario_lista
        
        
        
        """  # Crear una lista para almacenar la información del único usuario
        usuario_lista = []

        # Obtener los datos del único usuario
        for fila in hoja_reporte_usuarios.iter_rows(min_row=2, values_only=True):
            identificacion = fila[3]  # Suponiendo que la identificación está en la cuarta columna (índice 3)
            cargo = fila[2]  # Suponiendo que el cargo está en la tercera columna (índice 2)
            oficina = fila[1]  # Suponiendo que la oficina está en la segunda columna (índice 1)
            estado = fila[0]  # Suponiendo que el estado está en la primera columna (índice 0)
           
            if identificacion == num_identificacion:  # Comparar con el número de identificación que deseas buscar
                print(f"Coicidencia encontra para la identicacion: {num_identificacion}, Identificacion autorizada: {identificacion} ")
                usuario_lista = [cargo, oficina, estado,]
                return usuario_lista
        return None

        # Si el usuario existe, puedes acceder a los valores como:
        if usuario_lista:
            print(f"Identificación: {usuario_lista[0]}")
            print(f"Cargo: {usuario_lista[1]}")
            print(f"Estado: {usuario_lista[2]}")
            print(f"Oficina: {usuario_lista[3]}")
        else:
            print("Usuario no encontrado")
 """
    """ def guardar_en_excel(self, hoja):
        #Escribe los datos procesados en el archivo Excel en la hoja especificada.
        wb = self.excel.Workbooks.Open(self.path_ResultadoArchivo)
        try:
            ws = wb.Worksheets(hoja)
        except Exception as e:
            print(f"Error al acceder a la hoja '{hoja}': {e}")
            wb.Close(False)
            self.excel.Quit()
            return

        # Obtener la última fila con datos
        last_row = ws.Cells(ws.Rows.Count, 1).End(win32.constants.xlUp).Row + 1

         # Escribir datos según el mapeo
        print("Escribiendo datos en ADULTO MAYOR...")
        for i, row in self.reporte_pagos.iterrows():
            for col_letra, col_nombre in self.mapeo_columnas.items():
                col_idx = self.convertir_columna_a_indice(col_letra)
                valor = row[col_nombre] if not pd.isnull(row[col_nombre]) else ""
                ws.Cells(last_row + i, col_idx).Value = valor

            # Duplicar columna M (Oficina) en columna W
            col_idx_m = self.convertir_columna_a_indice('M')
            col_idx_w = self.convertir_columna_a_indice('W')
            ws.Cells(last_row + i, col_idx_w).Value = ws.Cells(last_row + i, col_idx_m).Value
            
        # Escribir datos de reporte_usuarios
        for i, row in self.reporte_usuarios.iterrows():
            for col_letra, col_nombre in self.mapeo_columnas_usuarios.items():
                col_idx = self.convertir_columna_a_indice(col_letra)
                valor = row[col_nombre] if not pd.isnull(row[col_nombre]) else ""
                ws.Cells(last_row + i, col_idx).Value = valor


        # Guardar y cerrar el archivo
        wb.Save()
        wb.Close()

    def cerrar_excel(self):
        Cierra la instancia de Excel.
        self.excel.Quit() """