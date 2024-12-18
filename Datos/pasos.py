import os
import datetime as dt
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from os.path import join, exists
from os import mkdir
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC
from Fuji.get_data import GetData
from Datos.logger import Logger

class Pasos:
    def __init__(self):
        load_dotenv()
        self.carpeta_nomPaso = None
        self.logger = None
        #rutas comunes en todas las clases
        self.path_ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        print("este es el path en la clase pasos ",self.path_)
        self.log=os.path.join(self.path_,'logger','logger.txt')
        self.logger=Logger(self.log)
        self.path_Reportes= os.path.join(self.path_,'Reportes')
        self.path_Resultado= os.path.join(self.path_,'Resultado')
        self.path_ResultadoArchivo= os.path.join(self.path_Resultado,'Adulto_mayor.xlsx')
        #self.db_siga_excel = os.path.join(self.path_, 'Insumo', 'plantillaParaActivacion1.xlsx')#db_proactivanet_siga
        self.__driver_path = os.path.join(self.path_, "driver_\\chromedriver-win64\\chromedriver.exe")
        # Construye la ruta dinámica
        self.download_dir = os.path.join(self.path_ ,'Reportes')
        #urls de saga
        self.SIGA_USUARIOS = os.getenv('Siga_Usuarios')

        #urls
        """ self.__siga_url = os.getenv('siga_url')
        self.siga_username = os.getenv('siga_username')
        self.siga_password = os.getenv('siga_password') """
        
        Data = GetData()
        __data = Data.get_datos_id('1')
        self.siga_username = __data['user_sig']
        self.siga_password = __data['pass_sig']
        self.__siga_url = __data['url']
       
    def crear_carpeta(self):
            try:
                 
                if not exists(self.path_Resultado):
                    mkdir(self.path_Resultado)

                if not exists(self.path_Reportes):
                    mkdir(self.path_Reportes)
                
                #self.carpeta_mes = join(self.path_Resultado, dt.datetime.today().strftime("%Y-%B-%d"))
                #if not exists(self.carpeta_mes):
                 #   mkdir(self.carpeta_mes)
                    

                #return log_file
            except Exception as e:
                self.logger.log(f"ERROR creando las carpetas: {e}")
                return None
            
    def __abrirPagina(self, siga_url):
        j = 0
        while j <2:
            try:
                # opciones de navegador, Iniciar las opciones de navegador (Chrome)
                # Iniciar el navegador con la ventana maximizada
                options = webdriver.ChromeOptions()
                options.add_argument('--start-maximized') 
                # desactivar todas las extensiones del navegador (Chrome)
                service = Service(executable_path=self.__driver_path)
                options.add_argument('--disable-extensions')
                prefs = {
                    "download.default_directory": self.download_dir,
                    "download.prompt_for_download": False,
                }
                options.add_experimental_option("prefs", prefs)
                driver = webdriver.Chrome(service=service, options=options)
                driver.get(siga_url)
                if not driver:
                    raise Exception("Error")
                self.logger.log(f"Pagina de siga iniciada correctamente en el paso:")
                return driver
            except Exception as e:
                j += 1
                self.logger.log(f"Error: iniciado la pagina de siga en el paso:  {e}")
                if j == 2:
                    return None
    
    def __login(self, driver):

        try:
            #Se ingresa el usuario
            driver.find_element(By.NAME, "login").send_keys(self.siga_username)
            # Ingresar la contraseña descifrada
            driver.find_element(By.ID,'password').send_keys(self.siga_password)
            #Click para ingresar
            btnAceptar = '//tbody/tr/td/a[@class="btn btn--primary"]'
            driver.find_element(By.XPATH, btnAceptar).click()
            time.sleep(2)
            self.logger.log(f"Sesion inicada correctamente en el paso: nomPaso")
            return True
        except Exception as e:
            self.logger.log(f"Error iniciando sesion en el paso  {e}")
            return False
    
    def iniciar_pagina(self):
        #Cargar la pagina de siga
        driver = self.__abrirPagina(self.__siga_url)
        if not self.esperar_elemento(driver, By.NAME, "login"):
            self.logger.log(f"No se pudo cargar la página de inicio de sesión ")
            return None
        #Iniciar sesion en siga
        conf= self.__login(driver)
        if not conf:
            self.logger.log(f"No se pudo obtener el driver en el paso ")
            return None
        return driver
    
    def esperar_elemento(self, driver, by_type, selector, timeout=5):
        try:
            element = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by_type, selector)))
            self.logger.log(f"Elemento cargado correctamente {selector}")
            return element
        except TimeoutException as e:
            self.logger.log(f"Tiempo agotado esperando el elemento {selector}: {e}")
            return None
    
    def leer_datos_excel(self, nom_hoja, ruta, nomPaso):
        try:
            df_users = pd.DataFrame()
            with open(join(self.path_, ruta), mode='rb') as fp:
                df_users = pd.read_excel(fp, sheet_name=f"{nom_hoja}", engine='openpyxl', dtype=str)
            self.logger.log(f"Datos de excel leidos con exito en el paso {nomPaso}")
            print(df_users)   
            return df_users
        except Exception as e:
            self.logger.log(f"Error al leer los datos de excel en el paso {nomPaso} {e}")   
            return
        
    def crear_df_saga(self, nomPaso):
        try:
            folder = join(self.carpeta_nomPaso, "claves_saga.xlsx")
            print(folder)
            if not os.path.exists(folder):
                # Crear un nuevo archivo Excel
                self.__wb = Workbook()
                hoja = self.__wb.active
                hoja.title = "SAGA"  # Nombrar la hoja activa
                hoja['A1'] = 'Identificacion'
                hoja['B1'] = 'Reingreso'
                self.__wb.save(folder)
            else:
                # Cargar el archivo existente
                self.__wb = load_workbook(folder)
            # Continuar trabajando con el archivo
            self.logger.log(f"df_saga creado con exito en el paso {nomPaso}") 
            return self.__wb
        except Exception as e:
            self.logger.log(f"Error creando el df_saga en el paso {nomPaso} {e}") 
            return
    
    def guardar_resultados_excel(self, wb, nomruta, guardados, no_creados):
        try:
            path_excel = join(self.carpeta_nomPaso, nomruta)
            """ if not exists(self.__path_historicos):
                mkdir(path_excel) """
            claves_saga = join(self.carpeta_nomPaso, "claves_saga.xlsx")
            df_guardados = pd.DataFrame(guardados)
            df_no_creados = pd.DataFrame(no_creados)
            
            with pd.ExcelWriter(path_excel) as writer:
                df_guardados.to_excel(writer, sheet_name="Usuarios_guardados", index=False)
                df_no_creados.to_excel(writer, sheet_name="Usuarios_No_Creados", index=False)
            wb.save(claves_saga)
            self.logger.log(f"Resultados guardados en {path_excel}\n\n")
        except Exception as e:
            self.logger.log(f"Error guardando los datos procesados en el excel{e}")
