import os
import zipfile
import pandas as pd
from datetime import datetime
import datetime as dt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re

class PasarDatos:
    def __init__(self):
        # Rutas
        self.path_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.path_matrix = os.path.join(self.path_root, 'reports')
        self.path_reports = os.path.join(self.path_root, 'reports_SIGA')
        # Dataframe global
        #self.df_aliados = pd.DataFrame()
    
    def cargar_archivos1(self):
        #Carga los archivos necesarios para el procesamiento.
        columnas_Pagos=['Fecha','Ciudad','Zona','Cédula Cajero/Vendedor','Cajero/Vendedor',
            'Canal', 'Reclama titular?', 'Empresa', 'Producto','Cod. Oficina', 'Oficina',
            'Cod. Sitio ','Sitio de venta','Tipo Doc. Titular', 'Identificación Titular',  'Titular',
            'Tipo Doc. Autorizado', 'Identificación Autorizado', 'Autorizado','Periodo', 'Valor Reportado', 
            'Valor Pagado', 'Valor Redondeo', 'Grupo Pago','Tipo Pago','Tipo Subsidio', 'Titular Comfama', 
            'Titular Comfama']
        self.reporte_pagos = pd.read_excel('reportes/reportePagos.xlsx',usecols=columnas_Pagos)
        print(self.reporte_pagos)

        columnas_Usuarios=['Identificacion','Cargo','Estado']
        self.reporte_usuarios = pd.read_excel('reportes/reporteUsuarios.xlsx',usecols=columnas_Usuarios)
        print(self.reporte_usuarios)

         #Procesa los datos cargados, separa fecha y hora, y selecciona las columnas de interés.
        # Separar la columna FechaHora en Fecha y Hora

        self.reporte_pagos.columns = self.reporte_pagos.columns.str.strip()
        self.reporte_pagos[['Fecha', 'Hora']] = self.reporte_pagos['Fecha'].str.split(' ', expand=True)
    
    def cargar_archivos(self):
        #Carga los archivos necesarios para el procesamiento.
        columnas_Pagos=['Fecha','Ciudad','Zona','Cédula Cajero/Vendedor','Cajero/Vendedor',
            'Canal', 'Reclama titular?', 'Empresa', 'Producto','Cod. Oficina', 'Oficina',
            'Cod. Sitio ','Sitio de venta','Tipo Doc. Titular', 'Identificación Titular',  'Titular',
            'Tipo Doc. Autorizado', 'Identificación Autorizado', 'Autorizado','Periodo', 'Valor Reportado', 
            'Valor Pagado', 'Valor Redondeo', 'Grupo Pago','Tipo Pago','Tipo Subsidio', 'Titular Comfama', 
            'Titular Comfama']
        self.reporte_pagos = load_workbook('reportes/reportePagos.xlsx')
        print(self.reporte_pagos)

        columnas_Usuarios=['Identificacion','Cargo','Estado']
        self.reporte_usuarios = load_workbook('reportes/reporteUsuarios.xlsx')
        print(self.reporte_usuarios)

         #Procesa los datos cargados, separa fecha y hora, y selecciona las columnas de interés.
        # Separar la columna FechaHora en Fecha y Hora

        self.reporte_pagos.columns = self.reporte_pagos.columns.str.strip()
        self.reporte_pagos[['Fecha', 'Hora']] = self.reporte_pagos['Fecha'].str.split(' ', expand=True)

    # Función para obtener la última fila no vacía
    def obtener_ultima_fila(self, df):
        ultima_fila = df.last_valid_index()  # Obtén el último índice válido (fila no vacía)
        return ultima_fila + 1 if ultima_fila is not None else 0

    # Recibe como parámetros la ruta del .xls y la ruta asignada para el .xlsx
    def convertir_xls_a_xlsx(self, path_xls, path_xlsx):
        # Leer el archivo .xls e indicar que la columna 'PIN GIRO' sea tipo texto
        data = pd.read_excel(path_xls, sheet_name=None, dtype={'PIN_GIRO': str}) # ,'PIN': str
        lista_nombres_M = ['Enviados M', 'Pagados M', 'Anulados M']
        # Se crea el archivo .xlsx en la ruta asignada
        with pd.ExcelWriter(path_xlsx, engine='openpyxl') as writer:
            # Se itera en las hojas del archivo .xls tomando el nombre para recrealas en el archivo .xlsx 
            for sheet_name, df in data.items():
                #print(df)
                if sheet_name == 'Anulados M':
                    df = self.validar(df)
                # Determinar la última fila no vacía
                ultima_fila = self.obtener_ultima_fila(df)
                # Filtrar filas válidas
                df = df.iloc[:ultima_fila].copy() if ultima_fila > 0 else df.copy() # Seleccionar filas desde el inicio hasta el índice 'ultimo' (excluye encabezado)
                # Realizar transformación para Hojas M 
                if sheet_name in lista_nombres_M:
                    df = self.transformar_datos(df)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"- Conversión de extesión de .xls a .xlsx realizada.")
        return path_xlsx
   
    # Recibe como parámetro el diccionario de archivos con su respectiva hoja y la ruta donde se encuentra el archivo
    def procesar_datos(self, archivos_hojas, matrix_path):
        # Cargar el archivo matriz
        wb_matrix = load_workbook(matrix_path)
    
        for archivo, hoja_destino in archivos_hojas.items():
            # Definir ruta del archivo de origen
            archivo_path = os.path.join(self.path_reports, archivo + ".xlsx")

            # Validar si el archivo de origen existe
            if os.path.exists(archivo_path):
                # Cargar el libro de con los datos de los reportes de SIGA
                wb_origen = load_workbook(archivo_path, data_only=True)

                # Iterar sobre las hojas del archivo origen
                for hoja_origen_nombre in wb_origen.sheetnames:
                    ws_origen = wb_origen[hoja_origen_nombre]

                    # Si la hoja de destino no existe, crearla
                    if hoja_destino not in wb_matrix.sheetnames:
                        ws_destino = wb_matrix.create_sheet(hoja_destino)
                    else:
                        ws_destino = wb_matrix[hoja_destino]

                    # Limpiar la hoja destino
                    ws_destino.delete_rows(1, ws_destino.max_row)

                    # Copiar datos de origen a destino
                    for i, fila in enumerate(ws_origen.iter_rows(values_only=True), start=1):
                        for j, valor in enumerate(fila, start=1):
                            # Aplicar transformaciones específicas
                            if isinstance(valor, dt.datetime) and 'Fecha' in (ws_origen.cell(1, j).value or ''):
                                try:
                                    # Convertir el valor datetime a string
                                    valor_str = valor.strftime('%H:%M:%S')  # Formato de hora HH:MM:SS
                                    # Aplicar la expresión regular al string
                                    pattern = r'(\d{1,2}:\d{2}:\d{2})'
                                    match = re.search(pattern, valor_str)
                                    if match:
                                        valor = match.group()  # Extraer el valor coincidente
                                    else:
                                        valor = None  # Manejar casos donde no hay coincidencia
                                except Exception as e:
                                    print(f"Error procesando valor: {valor}, Error: {e}")

                            if isinstance(valor, float) and 'Valor Giro' in (ws_origen.cell(1, j).value or ''):
                                valor = self.convertir_a_string(valor)
                            elif isinstance(valor, str) and 'Valor Giro' in (ws_origen.cell(1, j).value or ''):
                                valor = self.convertir_a_string(valor)
                            

                            # Escribir en la celda destino
                            ws_destino.cell(row=i, column=j, value=valor)

                print(f"- Datos procesados desde {archivo_path} hacia la hoja '{hoja_destino}' en la matriz.")

        # Guardar los cambios en el archivo matriz
        wb_matrix.save(matrix_path)
        print("- Procesamiento de datos realizado.")

    def obtener_dataframes(self, wb):
        # Crear un diccionario para almacenar las hojas como DataFrames
        hojas_df = {}

        # Iterar sobre todas las hojas en el workbook
        for hoja_nombre in wb.sheetnames:
            hoja = wb[hoja_nombre]  # Seleccionar la hoja
            datos = hoja.values
        
            # La primera fila como encabezados
            columnas = next(datos)
            # Crear el DataFrame
            df = pd.DataFrame(datos, columns=columnas)

            # Agregar el DataFrame al diccionario
            hojas_df[hoja_nombre] = df
        
        return hojas_df
    # Metodo principal
    def pasar_datos(self):
        try:
            print("Iniciando transferencia de datos de Siga a reporte aliado...")
            # Instanciar método de obtener hora
            hora = self.obtener_hora_actual()
            # Obtener fecha y hora actual
            fecha_hora_actual = f"{dt.datetime.now().day}-{dt.datetime.now().strftime('%m-%Y')}-{hora}"
            fecha_actual_zip = dt.datetime.now().strftime(f"%d-%m-%Y-{hora}")

            # Nombrar archivos
            matrix_filename = f"GANA {hora}.xls"
            zip_filename = f"Reportes_Giros_{fecha_actual_zip}.zip"

            # Definir rutas
            zip_path = os.path.join(self.path_reports, zip_filename)
            matrix_path = os.path.join(self.path_matrix, matrix_filename)
        
            # Validar sí hay archivos con extesión .xls para llamar al método convertir_xls_a_xlsx(
            if matrix_path.endswith('.xls'):
                matrix_path = self.convertir_xls_a_xlsx(matrix_path, matrix_path.replace('.xls', '.xlsx'))

            # Validar sí existe un archivo .zip para extraer los archivos 
            if os.path.exists(zip_path) and os.path.exists(matrix_path):
                with zipfile.ZipFile(zip_path, 'r') as zip_rep:
                    # validar si el .zip esta vacío
                    if len(zip_rep.namelist()) > 0:
                        zip_rep.extractall(self.path_reports)
                        nombres_zip = zip_rep.namelist()
                        #print(nombres_zip)
                    else:
                        mensaje ="El archivo ZIP está vacío."
                        raise FileNotFoundError(mensaje)

            # Diccionario con nombre de los archivos y sus hojas correspondientes
            archivos_hojas = {
                f"reporteGirosEnviados{fecha_hora_actual}": "Enviados R",
                f"reporteGirosPagados{fecha_hora_actual}": "Pagados R",
                f"reporteGirosAnulados{fecha_hora_actual}": "Anulados R"
            }

            # Validar que los nombres de los archivos dentro del .zip tengan el formato correcto
            self.validar_nombres_archivos(nombres_zip, archivos_hojas)

            # Llamar método procesar datos para llevar a cabo la transferencia enviando le los parámetros necesarios
            self.procesar_datos(archivos_hojas, matrix_path)

            # Libro de trabajo que contiene el archivo .xlsx
            wb_matrix = load_workbook(matrix_path)
            # Se llama el método para ocultar las columnas
            self.ocultar_columnas(wb_matrix, archivos_hojas.values())
            #dataframes = self.obtener_dataframes(wb_matrix)
            # Se guardan los cambios y procesos hechos
            wb_matrix.save(matrix_path)
            # Se llama método para eliminar archivos sueltos
            self.limpiar_archivos_temporales()

            print("Proceso de tranferencia de datos completado con éxito.")
            return True, "Se ejecutó correctamente."

        except Exception as e:
            mensaje_error = f"Error al procesar datos: {e}"
            print(mensaje_error)
            return False, mensaje_error
    
